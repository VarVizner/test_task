from django.shortcuts import render, get_object_or_404
from openpyxl import load_workbook
from .models import Elements
from .serializers import ElementsSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework import status

def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for el in ws.iter_rows(min_row=2, values_only=True):
            section, category, code, name, price = el
            if not Elements.objects.filter(code=code).exists():
                Elements.objects.create(section=section, category=category,
                                        code=code, name=name,
                                        price=price)

        return render(request, "success.html")

    return render(request, 'create.html')


@api_view(['GET'])
def ApiOverview(request):
    api_urls = {
        'Upload_excel_file': '/upload_excel',
        'All_items': '/all',
        'Add': '/add',
        'Update': '/update/code',
        'Delete': '/delete/code'
    }

    return Response(api_urls)


@api_view(['POST'])
def add_element(request):
    element = ElementsSerializer(data=request.data)

    if not element.is_valid():
        return Response(status=status.HTTP_400_BAD_REQUEST)

    if Elements.objects.filter(**request.data).exists():
        raise serializers.ValidationError('This element already exists')

    element.save()
    return Response(element.data)



@api_view(['GET'])
def view_elements(request):

    if request.query_params:
        element = Elements.objects.filter(**request.query_params.dict())
    else:
        element = Elements.objects.all()

    if element:
        serializer = ElementsSerializer(element, many=True)
        return Response(serializer.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['GET', 'PUT'])
def update_element(request, code):
    element = get_object_or_404(Elements, code=code)
    data = ElementsSerializer(instance=element, data=request.data)

    if data.is_valid():
        data.save()
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
def delete_element(request, code):
    element = get_object_or_404(Elements, code=code)
    element.delete()
    return Response(status=status.HTTP_200_OK)












